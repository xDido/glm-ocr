"""Generate a formula-driven Excel workbook for the AWS cost comparison.

Every USD in the sheet is a live `=cell*cell` formula. Change any of
the inputs at the top — hourly rates, storage sizes, hours/month,
NAT data volume — and the detailed breakdown and the pricing-tier
summary recalculate automatically when Excel opens/re-opens the file.

Prices: ap-southeast-1 (Singapore), Linux, Shared tenancy, pulled from
AWS's authoritative CSV:
  https://pricing.us-east-1.amazonaws.com/offers/v1.0/aws/AmazonEC2/current/ap-southeast-1/index.csv

On-demand / RI rates are AWS list prices for `RunInstances` (plain
Linux, no license bundle). RI columns are "No Upfront" for direct
hourly comparability. Spot is an estimate at typical Singapore-AZ
discount ratios (GPU ~45% off OD; CPU ~60% off OD).

Storage note: "all-NVMe" policy on the GPU host. g4dn.2xlarge includes
225 GB NVMe SSD (ephemeral) free. We mount it at /var/lib/docker AND
expose it at /root/.cache/huggingface inside the containers, so both
Docker image layers AND SGLang model weights live on NVMe. The GPU
host's EBS is the 30 GiB root only (AMI + system). Weights re-download
on every instance replacement (~2-5 min, ~$0.11 NAT-egress per replace).

CPU-side hosts (m7i / c7i in Options B/C) have NO included NVMe — they
run everything on a single 50 GiB root EBS. The ONNX cache (~140 MiB)
lives on root and re-exports in ~45 s on first boot after replacement.

Excel LTR-forced via sheet_view.rightToLeft = False.
"""
from __future__ import annotations

import pathlib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = pathlib.Path("loadtest/reports/aws-cost-singapore-1replica.xlsx")


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Singapore 1-replica"
    ws.sheet_view.rightToLeft = False

    # Styles
    title_font   = Font(bold=True, size=14, color="1F3864")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="2F5597")
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="4472C4")
    input_fill   = PatternFill("solid", fgColor="FFF2CC")   # soft yellow: "edit me"
    formula_fill = PatternFill("solid", fgColor="E2EFDA")   # soft green: computed
    total_font   = Font(bold=True)
    total_fill   = PatternFill("solid", fgColor="D9E1F2")
    cheapest_fill = PatternFill("solid", fgColor="C6EFCE")
    delta_plus   = PatternFill("solid", fgColor="FFC7CE")
    delta_minus  = PatternFill("solid", fgColor="C6EFCE")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align  = Alignment(horizontal="right", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")

    usd_fmt   = '"$"#,##0.00'
    rate_fmt  = '"$"0.0000'
    int_fmt   = '#,##0'
    plus_fmt  = '"+$"#,##0.00;"-$"#,##0.00'

    # Track current row cursor
    r = 1

    # === TITLE ===
    ws.cell(row=r, column=1, value="GLM-OCR deployment cost — Singapore (ap-southeast-1) · 1 replica").font = title_font
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2

    # === SECTION: Hourly rates (editable) ===
    sec_row = r
    cell = ws.cell(row=r, column=1, value="Hourly rates (edit yellow cells to recalculate)")
    cell.font = section_font; cell.fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    # Header row
    for c, h in enumerate(["Instance", "On-demand $/hr", "1yr RI NU $/hr", "3yr RI NU $/hr", "Spot (est) $/hr"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    r += 1

    # Rate inputs — remember row refs for later formulas.
    # g4dn, m7i, c7i rows
    RATE_ROWS = {}
    for name, od, ri1, ri3, spot in [
        ("g4dn.2xlarge", 1.0520, 0.6630, 0.5110, 0.5786),
        ("m7i.2xlarge",  0.5040, 0.3334, 0.2286, 0.2016),
        ("c7i.2xlarge",  0.4116, 0.2732, 0.1867, 0.1646),
    ]:
        ws.cell(row=r, column=1, value=name)
        for col, val in enumerate([od, ri1, ri3, spot], start=2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = rate_fmt
            cell.fill = input_fill
            cell.alignment = right_align
        RATE_ROWS[name] = r
        r += 1
    r += 1

    # === SECTION: Constants (editable) ===
    cell = ws.cell(row=r, column=1, value="Constants (editable — hours, per-unit rates, data volumes)")
    cell.font = section_font; cell.fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    CONST = {}
    for key, label, val, fmt in [
        ("hours",        "Hours per month",                             730,    int_fmt),
        ("gp3_rate",     "EBS gp3 ($/GiB-mo)",                          0.096,  rate_fmt),
        ("alb_hr",       "ALB hourly ($/hr)",                           0.0275, rate_fmt),
        ("alb_lcu_rate", "ALB LCU ($/LCU-hr)",                          0.008,  rate_fmt),
        ("alb_lcu_qty",  "LCU-hours consumed per month",                1000,   int_fmt),
        ("nat_hr",       "NAT Gateway hourly ($/hr)",                   0.059,  rate_fmt),
        ("nat_gb_rate",  "NAT data processed ($/GB)",                   0.059,  rate_fmt),
        ("nat_gb_a",     "NAT GB/mo — Option A",                        75,     int_fmt),
        ("nat_gb_bc",    "NAT GB/mo — Option B/C",                      80,     int_fmt),
        ("cw_rate",      "CloudWatch Logs ingest ($/GB)",               0.67,   rate_fmt),
        ("cw_gb_a",      "CW Logs GB/mo — Option A",                    5,      int_fmt),
        ("cw_gb_bc",     "CW Logs GB/mo — Option B/C",                  8,      int_fmt),
        ("secret_rate",  "Secrets Manager ($/secret-mo)",               0.40,   rate_fmt),
        ("secret_qty",   "Secrets count",                               1,      int_fmt),
        ("ecr_rate",     "ECR storage ($/GiB-mo)",                      0.10,   rate_fmt),
        ("ecr_qty",      "ECR storage GiB",                             3,      int_fmt),
    ]:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        cell.fill = input_fill
        cell.alignment = right_align
        CONST[key] = (r, 2)  # (row, col) of value cell
        r += 1
    r += 1

    # === SECTION: Storage sizes (editable) ===
    cell = ws.cell(row=r, column=1, value="Storage sizes — GPU host uses included NVMe for Docker graph; EBS keeps HF cache persistent")
    cell.font = section_font; cell.fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    STORAGE = {}
    for key, label, val in [
        ("root_gpu",   "Root EBS GiB per GPU host (30 + 225 GB NVMe for Docker + HF cache)", 30),
        ("root_cpu",   "Root EBS GiB per CPU host (m7i/c7i, no NVMe)",                       50),
        ("hf_gpu",     "HF cache EBS GiB — GPU side (0 = weights on NVMe, re-download on replace)", 0),
        ("hf_cpu",     "HF cache EBS GiB — CPU side (0 = ONNX on root, re-export on replace)",      0),
        ("hf_sidecar", "HF cache EBS GiB — sidecar (0 = everything on NVMe)",                0),
    ]:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = int_fmt
        cell.fill = input_fill
        cell.alignment = right_align
        STORAGE[key] = (r, 2)
        r += 1
    r += 1

    # Helpers to emit formula strings referencing input cells
    def cref(row_col):
        row, col = row_col
        return f"${get_column_letter(col)}${row}"

    hours_ref    = cref(CONST["hours"])
    gp3_ref      = cref(CONST["gp3_rate"])
    alb_hr_ref   = cref(CONST["alb_hr"])
    alb_lcu_rate = cref(CONST["alb_lcu_rate"])
    alb_lcu_qty  = cref(CONST["alb_lcu_qty"])
    nat_hr_ref   = cref(CONST["nat_hr"])
    nat_gb_rate  = cref(CONST["nat_gb_rate"])
    nat_gb_a     = cref(CONST["nat_gb_a"])
    nat_gb_bc    = cref(CONST["nat_gb_bc"])
    cw_rate      = cref(CONST["cw_rate"])
    cw_gb_a      = cref(CONST["cw_gb_a"])
    cw_gb_bc     = cref(CONST["cw_gb_bc"])
    secret_rate  = cref(CONST["secret_rate"])
    secret_qty   = cref(CONST["secret_qty"])
    ecr_rate     = cref(CONST["ecr_rate"])
    ecr_qty      = cref(CONST["ecr_qty"])
    root_gpu     = cref(STORAGE["root_gpu"])
    root_cpu     = cref(STORAGE["root_cpu"])
    hf_gpu       = cref(STORAGE["hf_gpu"])
    hf_cpu       = cref(STORAGE["hf_cpu"])
    hf_sidecar   = cref(STORAGE["hf_sidecar"])

    # Rate cell refs (column B for OD, C for 1yr, D for 3yr, E for spot)
    g4dn_od  = f"$B${RATE_ROWS['g4dn.2xlarge']}"
    g4dn_1yr = f"$C${RATE_ROWS['g4dn.2xlarge']}"
    g4dn_3yr = f"$D${RATE_ROWS['g4dn.2xlarge']}"
    g4dn_spot= f"$E${RATE_ROWS['g4dn.2xlarge']}"
    m7i_od   = f"$B${RATE_ROWS['m7i.2xlarge']}"
    m7i_1yr  = f"$C${RATE_ROWS['m7i.2xlarge']}"
    m7i_3yr  = f"$D${RATE_ROWS['m7i.2xlarge']}"
    m7i_spot = f"$E${RATE_ROWS['m7i.2xlarge']}"
    c7i_od   = f"$B${RATE_ROWS['c7i.2xlarge']}"
    c7i_1yr  = f"$C${RATE_ROWS['c7i.2xlarge']}"
    c7i_3yr  = f"$D${RATE_ROWS['c7i.2xlarge']}"
    c7i_spot = f"$E${RATE_ROWS['c7i.2xlarge']}"

    # === SECTION: Detailed breakdown (on-demand) ===
    cell = ws.cell(row=r, column=1, value="Detailed breakdown (on-demand) — USD columns are live formulas")
    cell.font = section_font; cell.fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    # Headers
    for c, h in enumerate([
        "Line item", "Rate reference",
        "A: Sidecar (qty)", "A: Sidecar (USD)",
        "B: g4dn + m7i (qty)", "B: g4dn + m7i (USD)",
        "C: g4dn + c7i (qty)", "C: g4dn + c7i (USD)",
    ], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    r += 1
    detail_first_row = r

    # Each row: (label, rate-ref text, a_qty_text, a_formula, b_qty_text, b_formula, c_qty_text, c_formula)
    # Formulas reference input cells so they recompute when inputs change.
    LINE_ITEMS = [
        ("EC2: g4dn.2xlarge (on-demand)",  f"={g4dn_od}&\"/hr\"", f"={hours_ref}&\" hr\"", f"={g4dn_od}*{hours_ref}",
                                                                   f"={hours_ref}&\" hr\"", f"={g4dn_od}*{hours_ref}",
                                                                   f"={hours_ref}&\" hr\"", f"={g4dn_od}*{hours_ref}"),
        ("EC2: m7i.2xlarge (on-demand)",   f"={m7i_od}&\"/hr\"",   "0 hr",                  "=0",
                                                                   f"={hours_ref}&\" hr\"", f"={m7i_od}*{hours_ref}",
                                                                   "0 hr",                  "=0"),
        ("EC2: c7i.2xlarge (on-demand)",   f"={c7i_od}&\"/hr\"",   "0 hr",                  "=0",
                                                                   "0 hr",                  "=0",
                                                                   f"={hours_ref}&\" hr\"", f"={c7i_od}*{hours_ref}"),
        ("Root EBS gp3",                   f"={gp3_ref}&\"/GiB-mo\"",
                                           f"={root_gpu}&\" GiB\"",                  f"={root_gpu}*{gp3_ref}",
                                           f"=({root_gpu}+{root_cpu})&\" GiB\"",     f"=({root_gpu}+{root_cpu})*{gp3_ref}",
                                           f"=({root_gpu}+{root_cpu})&\" GiB\"",     f"=({root_gpu}+{root_cpu})*{gp3_ref}"),
        ("HF cache EBS gp3 (GPU-side)",    f"={gp3_ref}&\"/GiB-mo\"",
                                           f"={hf_sidecar}&\" GiB\"",                f"={hf_sidecar}*{gp3_ref}",
                                           f"={hf_gpu}&\" GiB\"",                    f"={hf_gpu}*{gp3_ref}",
                                           f"={hf_gpu}&\" GiB\"",                    f"={hf_gpu}*{gp3_ref}"),
        ("HF cache EBS gp3 (CPU-side)",    f"={gp3_ref}&\"/GiB-mo\"",
                                           "0 GiB",                                  "=0",
                                           f"={hf_cpu}&\" GiB\"",                    f"={hf_cpu}*{gp3_ref}",
                                           f"={hf_cpu}&\" GiB\"",                    f"={hf_cpu}*{gp3_ref}"),
        ("Instance NVMe SSD (g4dn included, 225 GB, ephemeral)", "included in instance price",
                                           "225 GB",                                 "=0",
                                           "225 GB",                                 "=0",
                                           "225 GB",                                 "=0"),
        ("ALB hourly",                     f"={alb_hr_ref}&\"/hr\"",
                                           f"={hours_ref}&\" hr\"",                  f"={alb_hr_ref}*{hours_ref}",
                                           f"={hours_ref}&\" hr\"",                  f"={alb_hr_ref}*{hours_ref}",
                                           f"={hours_ref}&\" hr\"",                  f"={alb_hr_ref}*{hours_ref}"),
        ("ALB LCU",                        f"={alb_lcu_rate}&\"/LCU-hr\"",
                                           f"={alb_lcu_qty}&\" LCU-hr\"",            f"={alb_lcu_rate}*{alb_lcu_qty}",
                                           f"={alb_lcu_qty}&\" LCU-hr\"",            f"={alb_lcu_rate}*{alb_lcu_qty}",
                                           f"={alb_lcu_qty}&\" LCU-hr\"",            f"={alb_lcu_rate}*{alb_lcu_qty}"),
        ("NAT Gateway hourly",             f"={nat_hr_ref}&\"/hr\"",
                                           f"={hours_ref}&\" hr\"",                  f"={nat_hr_ref}*{hours_ref}",
                                           f"={hours_ref}&\" hr\"",                  f"={nat_hr_ref}*{hours_ref}",
                                           f"={hours_ref}&\" hr\"",                  f"={nat_hr_ref}*{hours_ref}"),
        ("NAT Gateway data processed",     f"={nat_gb_rate}&\"/GB\"",
                                           f"={nat_gb_a}&\" GB\"",                   f"={nat_gb_rate}*{nat_gb_a}",
                                           f"={nat_gb_bc}&\" GB\"",                  f"={nat_gb_rate}*{nat_gb_bc}",
                                           f"={nat_gb_bc}&\" GB\"",                  f"={nat_gb_rate}*{nat_gb_bc}"),
        ("CloudWatch Logs ingest",         f"={cw_rate}&\"/GB\"",
                                           f"={cw_gb_a}&\" GB\"",                    f"={cw_rate}*{cw_gb_a}",
                                           f"={cw_gb_bc}&\" GB\"",                   f"={cw_rate}*{cw_gb_bc}",
                                           f"={cw_gb_bc}&\" GB\"",                   f"={cw_rate}*{cw_gb_bc}"),
        ("Secrets Manager",                f"={secret_rate}&\"/secret-mo\"",
                                           f"={secret_qty}&\" secret\"",             f"={secret_rate}*{secret_qty}",
                                           f"={secret_qty}&\" secret\"",             f"={secret_rate}*{secret_qty}",
                                           f"={secret_qty}&\" secret\"",             f"={secret_rate}*{secret_qty}"),
        ("ECR storage",                    f"={ecr_rate}&\"/GiB-mo\"",
                                           f"={ecr_qty}&\" GiB\"",                   f"={ecr_rate}*{ecr_qty}",
                                           f"={ecr_qty}&\" GiB\"",                   f"={ecr_rate}*{ecr_qty}",
                                           f"={ecr_qty}&\" GiB\"",                   f"={ecr_rate}*{ecr_qty}"),
        ("Inter-container data transfer",  "$0/GB (loopback or same-AZ)",
                                           "0 GB",                                   "=0",
                                           "0 GB",                                   "=0",
                                           "0 GB",                                   "=0"),
    ]

    for item in LINE_ITEMS:
        label, rate_ref, a_qty, a_formula, b_qty, b_formula, c_qty, c_formula = item
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=rate_ref).alignment = right_align
        ws.cell(row=r, column=3, value=a_qty).alignment    = right_align
        cell = ws.cell(row=r, column=4, value=a_formula)
        cell.number_format = usd_fmt; cell.fill = formula_fill; cell.alignment = right_align
        ws.cell(row=r, column=5, value=b_qty).alignment    = right_align
        cell = ws.cell(row=r, column=6, value=b_formula)
        cell.number_format = usd_fmt; cell.fill = formula_fill; cell.alignment = right_align
        ws.cell(row=r, column=7, value=c_qty).alignment    = right_align
        cell = ws.cell(row=r, column=8, value=c_formula)
        cell.number_format = usd_fmt; cell.fill = formula_fill; cell.alignment = right_align
        r += 1
    detail_last_row = r - 1

    # Monthly total row (sum of each USD column)
    r += 1
    ws.cell(row=r, column=1, value="Monthly total (on-demand)").font = total_font
    for col in (4, 6, 8):
        letter = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"=SUM({letter}{detail_first_row}:{letter}{detail_last_row})")
        cell.number_format = usd_fmt; cell.fill = total_fill; cell.font = total_font
        cell.alignment = right_align
    TOTAL_OD_ROW = r
    r += 2

    # === SECTION: Non-compute subtotal (so we can subtract to get pure-compute) ===
    # Non-compute = monthly total minus the first 3 EC2 rows.
    ws.cell(row=r, column=1, value="Non-compute subtotal (everything except EC2 rows)").font = total_font
    for col in (4, 6, 8):
        letter = get_column_letter(col)
        # Non-compute = SUM(all) - SUM(first 3 EC2 rows)
        cell = ws.cell(row=r, column=col,
                       value=f"={letter}{TOTAL_OD_ROW}-SUM({letter}{detail_first_row}:{letter}{detail_first_row+2})")
        cell.number_format = usd_fmt; cell.fill = total_fill; cell.font = total_font
        cell.alignment = right_align
    NONCOMP_ROW = r
    r += 2

    # === SECTION: Pricing tier summary ===
    cell = ws.cell(row=r, column=1, value="Pricing tier summary (monthly totals; non-compute costs same across tiers)")
    cell.font = section_font; cell.fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for c, h in enumerate(
        ["Option", "On-demand", "1yr RI (No Upfront)", "3yr RI (No Upfront)",
         "3yr RI GPU + Spot CPU"], start=1
    ):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    r += 1

    # Each option: compute = sum of its EC2 rates × hours; total = compute + non-compute subtotal
    # For option A, non-compute uses column D; B uses F; C uses H.
    nc_a = f"$D${NONCOMP_ROW}"
    nc_b = f"$F${NONCOMP_ROW}"
    nc_c = f"$H${NONCOMP_ROW}"

    tier_rows_data = [
        ("A: Sidecar (g4dn only)",
         f"={g4dn_od}*{hours_ref}+{nc_a}",
         f"={g4dn_1yr}*{hours_ref}+{nc_a}",
         f"={g4dn_3yr}*{hours_ref}+{nc_a}",
         None),  # no CPU to spot
        ("B: Split g4dn + m7i",
         f"=({g4dn_od}+{m7i_od})*{hours_ref}+{nc_b}",
         f"=({g4dn_1yr}+{m7i_1yr})*{hours_ref}+{nc_b}",
         f"=({g4dn_3yr}+{m7i_3yr})*{hours_ref}+{nc_b}",
         f"=({g4dn_3yr}+{m7i_spot})*{hours_ref}+{nc_b}"),
        ("C: Split g4dn + c7i",
         f"=({g4dn_od}+{c7i_od})*{hours_ref}+{nc_c}",
         f"=({g4dn_1yr}+{c7i_1yr})*{hours_ref}+{nc_c}",
         f"=({g4dn_3yr}+{c7i_3yr})*{hours_ref}+{nc_c}",
         f"=({g4dn_3yr}+{c7i_spot})*{hours_ref}+{nc_c}"),
    ]
    tier_start = r
    for label, od_f, ri1_f, ri3_f, hyb_f in tier_rows_data:
        ws.cell(row=r, column=1, value=label).font = total_font
        for col, f in enumerate([od_f, ri1_f, ri3_f, hyb_f], start=2):
            if f is None:
                cell = ws.cell(row=r, column=col, value="N/A (no CPU to spot)")
                cell.font = Font(italic=True, color="808080")
                cell.alignment = right_align
            else:
                cell = ws.cell(row=r, column=col, value=f)
                cell.number_format = usd_fmt; cell.font = total_font
                cell.fill = formula_fill; cell.alignment = right_align
        r += 1
    tier_end = r - 1

    # Apply conditional formatting: highlight the minimum in each column green.
    from openpyxl.formatting.rule import FormulaRule
    for col in range(2, 6):
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{tier_start}:{col_letter}{tier_end}"
        # Rule: if this cell equals MIN of the numeric cells in the range, green-fill.
        rule = FormulaRule(
            formula=[f"AND(ISNUMBER({col_letter}{tier_start}), "
                     f"{col_letter}{tier_start}=MIN({rng}))"],
            fill=cheapest_fill,
        )
        # Need separate rules per row since "this row" isn't trivial in CF.
        # Simpler: set explicit rule for each cell in the column.
        for rr in range(tier_start, tier_end + 1):
            ws.conditional_formatting.add(
                f"{col_letter}{rr}",
                FormulaRule(
                    formula=[f"AND(ISNUMBER({col_letter}{rr}), "
                             f"{col_letter}{rr}=MIN({rng}))"],
                    fill=cheapest_fill,
                ),
            )
    r += 1

    # === SECTION: Deltas vs cheapest (Option A) ===
    cell = ws.cell(row=r, column=1, value="Delta vs Option A (cheapest) — positive = more expensive")
    cell.font = section_font; cell.fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for c, h in enumerate(["", "On-demand", "1yr RI NU", "3yr RI NU"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        if h:
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    r += 1

    # B - A and C - A for each of the three comparable tiers
    for label, row_offset in [("B − A", 1), ("C − A", 2)]:
        ws.cell(row=r, column=1, value=label).font = total_font
        for col, (a_col, other_col) in enumerate(
            [(2, 2), (3, 3), (4, 4)], start=2  # tier columns are 2/3/4 in the tier table
        ):
            a_cell = f"{get_column_letter(a_col)}{tier_start}"          # Option A row = tier_start
            other_cell = f"{get_column_letter(other_col)}{tier_start + row_offset}"
            cell = ws.cell(row=r, column=col, value=f"={other_cell}-{a_cell}")
            cell.number_format = plus_fmt
            cell.alignment = right_align
            cell.font = total_font
            # Tint via conditional formatting based on sign
            ws.conditional_formatting.add(
                cell.coordinate,
                FormulaRule(formula=[f"{cell.coordinate}>0"], fill=delta_plus),
            )
            ws.conditional_formatting.add(
                cell.coordinate,
                FormulaRule(formula=[f"{cell.coordinate}<0"], fill=delta_minus),
            )
        r += 1

    # === Column widths ===
    widths = [48, 22, 20, 20, 22, 22, 22, 22]
    for c_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
